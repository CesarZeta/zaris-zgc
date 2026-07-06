"""Maestro de artículos: CRUD, búsqueda, precios por lista (costo + margen),
cambio masivo de precios e importación desde Excel.

Regla de precios (legacy V16): precio_lista = costo_neto * (1 + utilidad/100),
donde costo_neto = costo / (1 + tasa_iva/100) si el costo cargado incluye IVA.
Si el artículo está en dólares, costo y precios se guardan en USD y se
convierten con la cotización vigente al mostrar/vender.
"""

import io
import uuid
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from typing import Literal

from fastapi import APIRouter, Depends, HTTPException, Response, UploadFile, status
from pydantic import BaseModel, Field
from sqlalchemy import Select, and_, func, or_, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.db import get_db
from app.core.permisos import requiere
from app.models import (
    Articulo,
    ArticuloStock,
    ArticuloVariante,
    Familia,
    Marca,
    Subfamilia,
    Unidad,
    Usuario,
)

router = APIRouter(prefix="/articulos", tags=["articulos"])

DOS = Decimal("0.01")
LISTAS = (1, 2, 3, 4)


# ===== Schemas =====

class ArticuloIn(BaseModel):
    codigo: str = Field(min_length=1, max_length=20)
    codigo_barras: str | None = Field(None, max_length=20)
    descripcion: str = Field(min_length=1, max_length=80)
    familia_id: uuid.UUID | None = None
    subfamilia_id: uuid.UUID | None = None
    marca_id: uuid.UUID | None = None
    unidad_id: uuid.UUID | None = None
    controla_stock: bool = True
    costo: Decimal = Field(Decimal("0"), ge=0)
    costo_con_iva: bool = False
    tasa_iva: Decimal = Field(Decimal("21"), ge=0, le=100)
    utilidad_1: Decimal | None = None
    utilidad_2: Decimal | None = None
    utilidad_3: Decimal | None = None
    utilidad_4: Decimal | None = None
    precio_1: Decimal | None = Field(None, ge=0)
    precio_2: Decimal | None = Field(None, ge=0)
    precio_3: Decimal | None = Field(None, ge=0)
    precio_4: Decimal | None = Field(None, ge=0)
    en_dolares: bool = False
    impuesto_interno: Decimal = Field(Decimal("0"), ge=0)
    pesable: bool = False
    venta_por_depto: bool = False
    es_envase_retornable: bool = False
    envase_articulo_id: uuid.UUID | None = None
    observaciones: str | None = None
    activo: bool = True


class ArticuloOut(BaseModel):
    id: uuid.UUID
    codigo: str
    codigo_barras: str | None
    descripcion: str
    familia_id: uuid.UUID | None
    subfamilia_id: uuid.UUID | None
    marca_id: uuid.UUID | None
    unidad_id: uuid.UUID | None
    controla_stock: bool
    costo: Decimal
    costo_con_iva: bool
    tasa_iva: Decimal
    utilidad_1: Decimal
    utilidad_2: Decimal
    utilidad_3: Decimal
    utilidad_4: Decimal
    precio_1: Decimal
    precio_2: Decimal
    precio_3: Decimal
    precio_4: Decimal
    en_dolares: bool
    impuesto_interno: Decimal
    pesable: bool
    venta_por_depto: bool
    es_envase_retornable: bool
    envase_articulo_id: uuid.UUID | None
    precio_actualizado_at: datetime | None
    observaciones: str | None
    activo: bool
    stock_total: Decimal | None = None

    model_config = {"from_attributes": True}


# ===== Lógica de precios =====

def costo_neto(costo: Decimal, costo_con_iva: bool, tasa_iva: Decimal) -> Decimal:
    if costo_con_iva and tasa_iva:
        return costo / (1 + tasa_iva / 100)
    return costo


def completar_precios(datos: dict, campos_enviados: set[str]) -> dict:
    """Completa la pata faltante de cada lista: si vino utilidad calcula el
    precio; si vino precio calcula la utilidad implícita. Legacy: editar
    cualquiera de las dos recalcula la otra."""
    neto = costo_neto(datos["costo"], datos["costo_con_iva"], datos["tasa_iva"])
    for i in LISTAS:
        u, p = f"utilidad_{i}", f"precio_{i}"
        util, precio = datos.get(u), datos.get(p)
        if p in campos_enviados and precio is not None:
            datos[p] = precio.quantize(DOS)
            if u not in campos_enviados:
                datos[u] = (
                    ((precio / neto - 1) * 100).quantize(DOS) if neto > 0 else Decimal("0")
                )
        elif util is not None:
            datos[u] = util
            datos[p] = (neto * (1 + util / 100)).quantize(DOS)
        else:
            datos.setdefault(u, Decimal("0"))
            datos.setdefault(p, Decimal("0"))
        if datos[u] is None:
            datos[u] = Decimal("0")
        if datos[p] is None:
            datos[p] = Decimal("0")
    return datos


CAMPOS_PRECIO = {"costo", "costo_con_iva", "tasa_iva"} | {
    f"{pre}_{i}" for pre in ("utilidad", "precio") for i in LISTAS
}


# ===== Helpers =====

def aplicar_busqueda_articulos(stmt: Select, q: str) -> Select:
    """Multi-palabra AND sobre descripción; los tokens también matchean
    código interno y código de barras (para lector en el maestro)."""
    q = q.strip()
    if not q:
        return stmt
    condiciones = []
    for token in q.split():
        patron = f"%{token}%"
        condiciones.append(
            or_(
                Articulo.descripcion.ilike(patron),
                Articulo.codigo.ilike(patron),
                Articulo.codigo_barras.ilike(patron),
            )
        )
    return stmt.where(and_(*condiciones))


async def _obtener_articulo(
    db: AsyncSession, tenant_id: uuid.UUID, articulo_id: uuid.UUID
) -> Articulo:
    articulo = await db.scalar(
        select(Articulo).where(Articulo.id == articulo_id, Articulo.tenant_id == tenant_id)
    )
    if articulo is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Artículo no encontrado")
    return articulo


async def _validar_referencias(db: AsyncSession, tenant_id: uuid.UUID, datos: dict) -> None:
    """Las FKs de catálogo deben pertenecer al tenant (y la subfamilia a su familia)."""
    checks = [
        ("familia_id", Familia),
        ("subfamilia_id", Subfamilia),
        ("marca_id", Marca),
        ("unidad_id", Unidad),
        ("envase_articulo_id", Articulo),
    ]
    for campo, modelo in checks:
        valor = datos.get(campo)
        if valor is None:
            continue
        obj = await db.scalar(
            select(modelo).where(modelo.id == valor, modelo.tenant_id == tenant_id)
        )
        if obj is None:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail=f"{campo}: no existe en la empresa",
            )
        if campo == "subfamilia_id" and datos.get("familia_id") not in (None, obj.familia_id):
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="La subfamilia no pertenece a la familia indicada",
            )


async def _validar_cbarra_contra_variantes(
    db: AsyncSession, tenant_id: uuid.UUID, cbarra: str | None
) -> None:
    """El EAN del artículo tampoco puede pisar el de una variante (unicidad
    cruzada entre tablas, a nivel app — espejo del chequeo en variantes.py)."""
    if not cbarra:
        return
    en_variante = await db.scalar(
        select(ArticuloVariante.id).where(
            ArticuloVariante.tenant_id == tenant_id,
            ArticuloVariante.codigo_barras == cbarra,
        )
    )
    if en_variante:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=f"El código de barras {cbarra} ya está en uso por una variante",
        )


def _subquery_stock_total(tenant_id: uuid.UUID):
    return (
        select(func.coalesce(func.sum(ArticuloStock.cantidad), 0))
        .where(
            ArticuloStock.articulo_id == Articulo.id,
            ArticuloStock.tenant_id == tenant_id,
        )
        .correlate(Articulo)
        .scalar_subquery()
    )


def _armar_out(articulo: Articulo, stock_total: Decimal | None) -> ArticuloOut:
    out = ArticuloOut.model_validate(articulo)
    out.stock_total = stock_total
    return out


# ===== CRUD =====

@router.post("", response_model=ArticuloOut, status_code=status.HTTP_201_CREATED)
async def crear_articulo(
    body: ArticuloIn,
    usuario: Usuario = Depends(requiere("articulos", "editar")),
    db: AsyncSession = Depends(get_db),
):
    datos = body.model_dump()
    datos["codigo"] = datos["codigo"].strip()
    if datos["codigo_barras"]:
        datos["codigo_barras"] = datos["codigo_barras"].strip() or None
    datos = completar_precios(datos, body.model_dump(exclude_unset=True).keys() | set())
    await _validar_referencias(db, usuario.tenant_id, datos)
    await _validar_cbarra_contra_variantes(db, usuario.tenant_id, datos["codigo_barras"])

    articulo = Articulo(
        tenant_id=usuario.tenant_id, precio_actualizado_at=datetime.now(timezone.utc), **datos
    )
    db.add(articulo)
    try:
        await db.commit()
    except IntegrityError:
        await db.rollback()
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Ya existe un artículo con ese código o código de barras",
        )
    return _armar_out(articulo, Decimal("0"))


@router.get("", response_model=list[ArticuloOut])
async def listar_articulos(
    response: Response,
    q: str = "",
    familia_id: uuid.UUID | None = None,
    subfamilia_id: uuid.UUID | None = None,
    marca_id: uuid.UUID | None = None,
    incluir_inactivos: bool = False,
    limit: int = 50,
    offset: int = 0,
    usuario: Usuario = Depends(requiere("articulos", "ver")),
    db: AsyncSession = Depends(get_db),
):
    stmt = select(Articulo).where(Articulo.tenant_id == usuario.tenant_id)
    if not incluir_inactivos:
        stmt = stmt.where(Articulo.activo)
    if familia_id:
        stmt = stmt.where(Articulo.familia_id == familia_id)
    if subfamilia_id:
        stmt = stmt.where(Articulo.subfamilia_id == subfamilia_id)
    if marca_id:
        stmt = stmt.where(Articulo.marca_id == marca_id)
    stmt = aplicar_busqueda_articulos(stmt, q)

    # El COUNT va sin la subquery de stock: sumar el stock del catálogo
    # entero solo para contar filas es O(stock) inútil.
    total = await db.scalar(select(func.count()).select_from(stmt.subquery()))
    response.headers["X-Total-Count"] = str(total or 0)

    stock_total = _subquery_stock_total(usuario.tenant_id)
    stmt = stmt.add_columns(stock_total.label("stock_total"))
    stmt = stmt.order_by(Articulo.descripcion).limit(min(limit, 200)).offset(offset)
    filas = (await db.execute(stmt)).all()
    return [_armar_out(articulo, st) for articulo, st in filas]


@router.get("/{articulo_id}", response_model=ArticuloOut)
async def obtener_articulo(
    articulo_id: uuid.UUID,
    usuario: Usuario = Depends(requiere("articulos", "ver")),
    db: AsyncSession = Depends(get_db),
):
    articulo = await _obtener_articulo(db, usuario.tenant_id, articulo_id)
    stock = await db.scalar(
        select(func.coalesce(func.sum(ArticuloStock.cantidad), 0)).where(
            ArticuloStock.articulo_id == articulo_id,
            ArticuloStock.tenant_id == usuario.tenant_id,
        )
    )
    return _armar_out(articulo, stock)


@router.put("/{articulo_id}", response_model=ArticuloOut)
async def actualizar_articulo(
    articulo_id: uuid.UUID,
    body: ArticuloIn,
    usuario: Usuario = Depends(requiere("articulos", "editar")),
    db: AsyncSession = Depends(get_db),
):
    articulo = await _obtener_articulo(db, usuario.tenant_id, articulo_id)

    enviados = set(body.model_dump(exclude_unset=True).keys())
    datos = body.model_dump()
    datos["codigo"] = datos["codigo"].strip()
    if datos["codigo_barras"]:
        datos["codigo_barras"] = datos["codigo_barras"].strip() or None
    datos = completar_precios(datos, enviados)
    await _validar_referencias(db, usuario.tenant_id, datos)
    if datos["codigo_barras"] != articulo.codigo_barras:
        await _validar_cbarra_contra_variantes(db, usuario.tenant_id, datos["codigo_barras"])

    toca_precios = any(
        getattr(articulo, campo) != datos[campo] for campo in CAMPOS_PRECIO
    )
    for campo, valor in datos.items():
        setattr(articulo, campo, valor)
    articulo.updated_at = func.now()
    if toca_precios:
        articulo.precio_actualizado_at = datetime.now(timezone.utc)

    try:
        await db.commit()
    except IntegrityError:
        await db.rollback()
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Ya existe un artículo con ese código o código de barras",
        )
    articulo = await db.scalar(select(Articulo).where(Articulo.id == articulo_id))
    stock = await db.scalar(
        select(func.coalesce(func.sum(ArticuloStock.cantidad), 0)).where(
            ArticuloStock.articulo_id == articulo_id,
            ArticuloStock.tenant_id == usuario.tenant_id,
        )
    )
    return _armar_out(articulo, stock)


# ===== Cambio masivo de precios =====

class CambioPreciosIn(BaseModel):
    tipo: Literal["porcentaje_precios", "porcentaje_costo", "fijar_margen"]
    porcentaje: Decimal
    listas: list[int] = Field(default=[1, 2, 3, 4])
    familia_id: uuid.UUID | None = None
    subfamilia_id: uuid.UUID | None = None
    marca_id: uuid.UUID | None = None
    en_dolares: bool | None = None
    q: str = ""
    dry_run: bool = False


class CambioPreciosOut(BaseModel):
    afectados: int
    dry_run: bool
    muestra: list[dict]


@router.post("/cambio-precios", response_model=CambioPreciosOut)
async def cambio_masivo_precios(
    body: CambioPreciosIn,
    usuario: Usuario = Depends(requiere("articulos", "editar")),
    db: AsyncSession = Depends(get_db),
):
    """Legacy 'Cambios masivos de precios': porcentaje sobre precios (mantiene
    costo, recalcula margen), porcentaje sobre costo (mantiene margen,
    recalcula precios) o fijar margen (utilidad = porcentaje en las listas
    elegidas). Con dry_run=true solo muestra el efecto."""
    listas = sorted({i for i in body.listas if i in LISTAS})
    if not listas:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Indicar al menos una lista 1-4"
        )
    if body.tipo != "fijar_margen" and body.porcentaje <= -100:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="El porcentaje debe ser mayor a -100",
        )

    stmt = select(Articulo).where(Articulo.tenant_id == usuario.tenant_id, Articulo.activo)
    if body.familia_id:
        stmt = stmt.where(Articulo.familia_id == body.familia_id)
    if body.subfamilia_id:
        stmt = stmt.where(Articulo.subfamilia_id == body.subfamilia_id)
    if body.marca_id:
        stmt = stmt.where(Articulo.marca_id == body.marca_id)
    if body.en_dolares is not None:
        stmt = stmt.where(Articulo.en_dolares == body.en_dolares)
    stmt = aplicar_busqueda_articulos(stmt, body.q)

    articulos = (await db.scalars(stmt)).all()
    factor = 1 + body.porcentaje / 100
    ahora = datetime.now(timezone.utc)
    muestra = []

    for art in articulos:
        antes = art.precio_1
        if body.tipo == "porcentaje_costo":
            art.costo = (art.costo * factor).quantize(Decimal("0.0001"))
        neto = costo_neto(art.costo, art.costo_con_iva, art.tasa_iva)
        for i in listas:
            u, p = f"utilidad_{i}", f"precio_{i}"
            if body.tipo == "porcentaje_precios":
                nuevo = (getattr(art, p) * factor).quantize(DOS)
                setattr(art, p, nuevo)
                if neto > 0:
                    setattr(art, u, ((nuevo / neto - 1) * 100).quantize(DOS))
            elif body.tipo == "porcentaje_costo":
                setattr(art, p, (neto * (1 + getattr(art, u) / 100)).quantize(DOS))
            else:  # fijar_margen
                setattr(art, u, body.porcentaje)
                setattr(art, p, (neto * factor).quantize(DOS))
        art.precio_actualizado_at = ahora
        art.updated_at = func.now()
        if len(muestra) < 5:
            muestra.append(
                {
                    "codigo": art.codigo,
                    "descripcion": art.descripcion,
                    "precio_1_antes": str(antes),
                    "precio_1_despues": str(art.precio_1),
                }
            )

    if body.dry_run:
        await db.rollback()
    else:
        await db.commit()
    return CambioPreciosOut(afectados=len(articulos), dry_run=body.dry_run, muestra=muestra)


# ===== Importación desde Excel =====

COLUMNAS_EXCEL = {
    "codigo", "codigo_barras", "descripcion", "familia", "subfamilia", "marca",
    "unidad", "costo", "costo_con_iva", "tasa_iva", "en_dolares",
    "utilidad_1", "utilidad_2", "utilidad_3", "utilidad_4",
    "precio_1", "precio_2", "precio_3", "precio_4",
}


class ImportExcelOut(BaseModel):
    total_filas: int
    creados: int
    actualizados: int
    errores: list[dict]


def _celda_decimal(valor, campo: str) -> Decimal | None:
    if valor is None or valor == "":
        return None
    try:
        return Decimal(str(valor).replace(",", "."))
    except InvalidOperation:
        raise ValueError(f"{campo}: '{valor}' no es un número")


def _celda_bool(valor) -> bool | None:
    if valor is None or valor == "":
        return None
    if isinstance(valor, bool):
        return valor
    return str(valor).strip().lower() in ("1", "si", "sí", "true", "x", "s")


@router.post("/importar-excel", response_model=ImportExcelOut)
async def importar_excel(
    archivo: UploadFile,
    usuario: Usuario = Depends(requiere("articulos", "editar")),
    db: AsyncSession = Depends(get_db),
):
    """Alta/actualización masiva desde .xlsx (feature heredada del legacy).
    Fila 1 = encabezados (codigo y descripcion obligatorios; familia, subfamilia,
    marca y unidad se crean si no existen). Upsert por código."""
    from openpyxl import load_workbook

    if not (archivo.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Subir un archivo .xlsx"
        )
    contenido = await archivo.read()
    try:
        wb = load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="No se pudo leer el .xlsx"
        )
    hoja = wb.active
    filas = hoja.iter_rows(values_only=True)
    try:
        encabezados = [str(c or "").strip().lower().replace(" ", "_") for c in next(filas)]
    except StopIteration:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Planilla vacía")
    if "codigo" not in encabezados or "descripcion" not in encabezados:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="La fila 1 debe tener al menos las columnas: codigo, descripcion",
        )
    desconocidas = [c for c in encabezados if c and c not in COLUMNAS_EXCEL]

    # Cachés de catálogos por nombre (se crean on-the-fly, como el legacy)
    tenant = usuario.tenant_id

    async def _catalogo(modelo, nombre: str, extra: dict | None = None):
        clave = nombre.strip()
        if not clave:
            return None
        obj = await db.scalar(
            select(modelo).where(modelo.tenant_id == tenant, modelo.nombre == clave[:40])
        )
        if obj is None:
            obj = modelo(tenant_id=tenant, nombre=clave[:40], **(extra or {}))
            db.add(obj)
            await db.flush()
        return obj

    creados = actualizados = total = 0
    errores: list[dict] = []
    if desconocidas:
        errores.append({"fila": 1, "error": f"Columnas ignoradas: {', '.join(desconocidas)}"})

    for nro, fila in enumerate(filas, start=2):
        if fila is None or all(c is None or str(c).strip() == "" for c in fila):
            continue
        total += 1
        celdas = dict(zip(encabezados, fila))
        try:
          # savepoint por fila: un error descarta solo esa fila, no lo ya importado
          async with db.begin_nested():
            codigo = str(celdas.get("codigo") or "").strip()
            descripcion = str(celdas.get("descripcion") or "").strip()
            if not codigo or not descripcion:
                raise ValueError("codigo y descripcion son obligatorios")

            datos: dict = {"codigo": codigo[:20], "descripcion": descripcion[:80]}
            if celdas.get("codigo_barras"):
                cb = str(celdas["codigo_barras"]).strip()
                # Excel suele traer códigos numéricos como float: 779123.0
                datos["codigo_barras"] = cb.removesuffix(".0")[:20] or None

            if celdas.get("familia"):
                familia = await _catalogo(Familia, str(celdas["familia"]))
                datos["familia_id"] = familia.id if familia else None
                if celdas.get("subfamilia") and familia:
                    nombre_sub = str(celdas["subfamilia"]).strip()[:40]
                    sub = await db.scalar(
                        select(Subfamilia).where(
                            Subfamilia.tenant_id == tenant,
                            Subfamilia.familia_id == familia.id,
                            Subfamilia.nombre == nombre_sub,
                        )
                    )
                    if sub is None and nombre_sub:
                        sub = Subfamilia(tenant_id=tenant, familia_id=familia.id, nombre=nombre_sub)
                        db.add(sub)
                        await db.flush()
                    datos["subfamilia_id"] = sub.id if sub else None
            if celdas.get("marca"):
                marca = await _catalogo(Marca, str(celdas["marca"]))
                datos["marca_id"] = marca.id if marca else None
            if celdas.get("unidad"):
                cod_uni = str(celdas["unidad"]).strip().upper()[:6]
                if cod_uni:
                    unidad = await db.scalar(
                        select(Unidad).where(Unidad.tenant_id == tenant, Unidad.codigo == cod_uni)
                    )
                    if unidad is None:
                        unidad = Unidad(tenant_id=tenant, codigo=cod_uni, nombre=cod_uni)
                        db.add(unidad)
                        await db.flush()
                    datos["unidad_id"] = unidad.id

            for campo in ("costo", "tasa_iva", "utilidad_1", "utilidad_2", "utilidad_3",
                          "utilidad_4", "precio_1", "precio_2", "precio_3", "precio_4"):
                valor = _celda_decimal(celdas.get(campo), campo)
                if valor is not None:
                    datos[campo] = valor
            for campo in ("costo_con_iva", "en_dolares"):
                valor = _celda_bool(celdas.get(campo))
                if valor is not None:
                    datos[campo] = valor

            existente = await db.scalar(
                select(Articulo).where(Articulo.tenant_id == tenant, Articulo.codigo == datos["codigo"])
            )
            if existente is None:
                base = {"costo": Decimal("0"), "costo_con_iva": False, "tasa_iva": Decimal("21")}
                base.update(datos)
                base = completar_precios(base, set(datos.keys()))
                db.add(
                    Articulo(
                        tenant_id=tenant,
                        precio_actualizado_at=datetime.now(timezone.utc),
                        **base,
                    )
                )
                creados += 1
            else:
                completos = {
                    "costo": datos.get("costo", existente.costo),
                    "costo_con_iva": datos.get("costo_con_iva", existente.costo_con_iva),
                    "tasa_iva": datos.get("tasa_iva", existente.tasa_iva),
                }
                for i in LISTAS:
                    completos[f"utilidad_{i}"] = datos.get(
                        f"utilidad_{i}", getattr(existente, f"utilidad_{i}")
                    )
                    if f"precio_{i}" in datos:
                        completos[f"precio_{i}"] = datos[f"precio_{i}"]
                completos = completar_precios(completos, set(datos.keys()))
                for campo, valor in {**datos, **completos}.items():
                    setattr(existente, campo, valor)
                existente.precio_actualizado_at = datetime.now(timezone.utc)
                existente.updated_at = func.now()
                actualizados += 1
            await db.flush()
        except (ValueError, IntegrityError) as e:
            # el savepoint ya deshizo la fila; el resto del import sigue vivo
            errores.append({"fila": nro, "error": str(e)})

    await db.commit()
    return ImportExcelOut(
        total_filas=total, creados=creados, actualizados=actualizados, errores=errores
    )
